"""
Export Family Tree to Excel.

Sheet 1 — แผนผังตระกูล
  Top-down tree with:
  - Left sidebar: generation number + year range (BE)
  - Colored node cards (dark = old generation, light = new)
  - Connector lines drawn with colored cells
  - Matches the web UI layout exactly

Sheet 2 — ข้อมูลสมาชิกทั้งหมด
  Full member table with all fields

Sheet 3 — วิธีใช้งาน
  Legend and instructions
"""
import builtins
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import list_members, get_roots, build_tree, get_conn
from date_utils import display_year, display_date

OUTPUT = Path(__file__).parent / "family_tree.xlsx"

# ─── Layout constants ─────────────────────────────────────────────────────────
NODE_W    = 5   # columns per person-card (including 1 padding on each side)
NODE_H    = 4   # rows per generation band
SIB_GAP   = 1   # gap columns between sibling subtrees
ROOT_GAP  = 3   # gap columns between root families
CONN_ROWS = 4   # connector rows between two generations
                # breakdown: 1 vert-down + 1 horiz + 1 vert-up + 1 spacer
SB_COL_GEN  = 1  # sidebar: generation label (col A)
SB_COL_YEAR = 2  # sidebar: year range (col B)
SB_GAP_COL  = 3  # gap column (col C)
TREE_START  = 4  # first tree column (col D, 1-indexed)
TITLE_ROWS  = 2  # rows for title + subtitle

# ─── Color palette ────────────────────────────────────────────────────────────
# (node_bg, node_fg, node_border)  — index = generation depth
GEN_PALETTE = [
    ('1A472A', 'FFFFFF', '0D2B18'),
    ('2D6A4F', 'FFFFFF', '1A472A'),
    ('40916C', 'FFFFFF', '2D6A4F'),
    ('52B788', '1A1A1A', '2D6A4F'),
    ('74C69D', '1A1A1A', '40916C'),
    ('95D5B2', '1A1A1A', '52B788'),
    ('B7E4C7', '1A1A1A', '74C69D'),
    ('D8F3DC', '1A1A1A', '95D5B2'),
]

CONN_COLOR   = '2D6A4F'   # connector line color
SB_GEN_FILL  = '1A1A2E'   # sidebar gen-label background
SB_YEAR_FILL = 'F4FBF6'   # sidebar year background
TITLE_FILL   = '1A1A2E'   # title row background
HEADER_FILL  = '2C3E50'   # table header background


def _fill(hex_c: str) -> PatternFill:
    # openpyxl requires 8-char ARGB; prepend FF for full opacity
    if len(hex_c) == 6:
        hex_c = 'FF' + hex_c
    return PatternFill('solid', fgColor=hex_c)


def _border(color='CCCCCC', style='thin') -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ─── Tree layout algorithm ────────────────────────────────────────────────────

def subtree_width(node: dict) -> int:
    """Minimum columns needed to display this subtree."""
    children = node.get('children', [])
    if not children:
        return NODE_W
    child_total = (sum(subtree_width(c) for c in children)
                   + SIB_GAP * (len(children) - 1))
    return max(NODE_W, child_total)


def assign_positions(node: dict, col_start: int, depth: int,
                     pos: dict) -> None:
    """
    Recursively assign {node_id: {depth, col, center, width}}.
    col_start / center are in *tree-local* column units (0-indexed).
    """
    children = node.get('children', [])
    w = subtree_width(node)

    if not children:
        center = col_start + (NODE_W - 1) // 2
        pos[node['id']] = {
            'depth': depth, 'col': col_start,
            'center': center, 'width': w,
        }
        return

    cur = col_start
    for child in children:
        cw = subtree_width(child)
        assign_positions(child, cur, depth + 1, pos)
        cur += cw + SIB_GAP

    fc = pos[children[0]['id']]['center']
    lc = pos[children[-1]['id']]['center']
    my_center = (fc + lc) // 2
    my_col = max(col_start, my_center - (NODE_W - 1) // 2)
    pos[node['id']] = {
        'depth': depth, 'col': my_col,
        'center': my_center, 'width': w,
    }


def tree_depth(node: dict) -> int:
    if not node.get('children'):
        return 0
    return 1 + max(tree_depth(c) for c in node['children'])


def row_for_depth(depth: int) -> int:
    """Excel row (1-indexed) where the node card starts for a given depth."""
    return TITLE_ROWS + depth * (NODE_H + CONN_ROWS) + 1


def col_for_tc(tc: int) -> int:
    """Excel column (1-indexed) for a 0-based tree column."""
    return TREE_START + tc


def collect_gen_years(node: dict, acc: dict) -> None:
    d = node['depth']
    for person in [node] + node.get('spouses', []):
        yr = display_year(person.get('birth_date') or '', short=True)
        if yr and yr.isdigit():
            acc.setdefault(d, []).append(int(yr))
    for child in node.get('children', []):
        collect_gen_years(child, acc)


def flatten_tree(node: dict, out: list) -> None:
    out.append(node)
    for c in node.get('children', []):
        flatten_tree(c, out)


# ─── Excel rendering helpers ──────────────────────────────────────────────────

def write_merged(ws, r1, c1, r2, c2, value, font, fill, align, border=None):
    """Merge a rectangle of cells, style + write the top-left cell."""
    if r1 != r2 or c1 != c2:
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1)
    cell.value = value
    cell.font = font
    cell.fill = fill
    cell.alignment = align
    if border:
        cell.border = border
    return cell


def write_node_card(ws, node: dict, p: dict) -> None:
    d = p['depth']
    er = row_for_depth(d)
    ec = col_for_tc(p['col'])

    fill_idx = min(d, len(GEN_PALETTE) - 1)
    bg, fg, bdr_c = GEN_PALETTE[fill_idx]

    # ── Build card text ──────────────────────────────────────────────────────
    name = f"{node['first_name']} {node['last_name']}"
    if node.get('nickname'):
        name += f" ({node['nickname']})"

    lines = [name]

    for sp in node.get('spouses', []):
        sp_name = f"{sp['first_name']} {sp['last_name']}"
        if sp.get('nickname'):
            sp_name += f" ({sp['nickname']})"
        lines.append(f"♥ {sp_name}")  # ♥

    birth_yr = display_year(node.get('birth_date') or '', short=True)
    death_yr = display_year(node.get('death_date') or '', short=True)
    if birth_yr:
        yr_str = f"พ.ศ. {birth_yr}" + (f" – {death_yr}" if death_yr else '')
        lines.append(yr_str)
    elif death_yr:
        lines.append(f"– พ.ศ. {death_yr}")
    elif (node.get('birth_date') or '').upper() == 'UNKNOWN':
        lines.append('ไม่พบข้อมูล')

    if node.get('occupation'):
        lines.append(node['occupation'])

    text = '\n'.join(lines)

    # ── Write merged cell ────────────────────────────────────────────────────
    write_merged(
        ws, er, ec, er + NODE_H - 1, ec + NODE_W - 1,
        value=text,
        font=Font(bold=True, color=fg, size=10, name='Sarabun'),
        fill=_fill(bg),
        align=Alignment(horizontal='center', vertical='center',
                        wrap_text=True),
        border=Border(
            left=Side(style='medium', color=bdr_c),
            right=Side(style='medium', color=bdr_c),
            top=Side(style='medium', color=bdr_c),
            bottom=Side(style='medium', color=bdr_c),
        ),
    )


def _color_cell(ws, row: int, col: int, color: str) -> None:
    """Color a single cell if it's empty (don't overwrite node cards)."""
    cell = ws.cell(row, col)
    if cell.value is None:
        cell.fill = _fill(color)


def draw_connectors(ws, node: dict, pos: dict) -> None:
    """Draw T-shaped connectors from a node down to its children."""
    children = node.get('children', [])
    if not children:
        return

    p = pos[node['id']]
    parent_last_row = row_for_depth(p['depth']) + NODE_H - 1
    parent_ec = col_for_tc(p['center'])

    # Rows in the connector band:
    # [parent_last_row + 1] vertical-down stub
    # [parent_last_row + 2] horizontal bar
    # [parent_last_row + 3] vertical-up stub to each child
    # [parent_last_row + 4] = child's first row
    cr_v = parent_last_row + 1   # vertical stub from parent
    cr_h = parent_last_row + 2   # horizontal bar
    cr_d = parent_last_row + 3   # vertical drops to children

    child_centers_ec = [col_for_tc(pos[c['id']]['center']) for c in children]
    min_ec = min(child_centers_ec)
    max_ec = max(child_centers_ec)

    # 1. Vertical down from parent
    _color_cell(ws, cr_v, parent_ec, CONN_COLOR)

    # 2. Horizontal bar (from leftmost child center to rightmost)
    for ec in range(min_ec, max_ec + 1):
        _color_cell(ws, cr_h, ec, CONN_COLOR)

    # 3. Vertical drops to each child
    for child_ec in child_centers_ec:
        _color_cell(ws, cr_d, child_ec, CONN_COLOR)

    for child in children:
        draw_connectors(ws, child, pos)


def write_sidebar_label(ws, d: int, gen_years: dict) -> None:
    """Write generation number + year range in the sidebar columns."""
    er = row_for_depth(d)

    # Gen number cell
    write_merged(
        ws, er, SB_COL_GEN, er + NODE_H - 1, SB_COL_GEN,
        value=f'รุ่น\n{d + 1}',
        font=Font(bold=True, size=11, color='FFFFFF', name='Sarabun'),
        fill=_fill(GEN_PALETTE[min(d, len(GEN_PALETTE) - 1)][0]),
        align=Alignment(horizontal='center', vertical='center', wrap_text=True),
    )

    # Year range cell
    years = gen_years.get(d, [])
    if years:
        mn, mx = min(years), max(years)
        yr_text = (f'พ.ศ.\n{mn}' if mn == mx
                   else f'พ.ศ.\n{mn}\n– {mx}')
    else:
        yr_text = ''
    write_merged(
        ws, er, SB_COL_YEAR, er + NODE_H - 1, SB_COL_YEAR,
        value=yr_text,
        font=Font(size=10, color='444444', name='Sarabun'),
        fill=_fill(SB_YEAR_FILL),
        align=Alignment(horizontal='center', vertical='center', wrap_text=True),
    )


# ─── Sheet 1: Tree ────────────────────────────────────────────────────────────

def build_sheet_tree(wb: openpyxl.Workbook) -> None:
    ws = wb.active
    ws.title = 'แผนผังตระกูล'
    ws.sheet_view.showGridLines = False

    roots = get_roots()
    if not roots:
        ws.cell(1, 1).value = 'ยังไม่มีข้อมูล'
        return

    trees = [build_tree(r['id']) for r in roots]

    # ── Layout: assign positions for all trees side-by-side ──────────────────
    all_pos: dict = {}
    col_cursor = 0
    for tree in trees:
        assign_positions(tree, col_cursor, 0, all_pos)
        col_cursor += subtree_width(tree) + ROOT_GAP
    total_tree_cols = max(0, col_cursor - ROOT_GAP)

    max_d = max((tree_depth(t) for t in trees), default=0)

    # Collect year ranges per generation
    gen_years: dict = {}
    for tree in trees:
        collect_gen_years(tree, gen_years)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(SB_COL_GEN)].width  = 7
    ws.column_dimensions[get_column_letter(SB_COL_YEAR)].width = 11
    ws.column_dimensions[get_column_letter(SB_GAP_COL)].width  = 1.5
    for i in range(total_tree_cols + ROOT_GAP + 2):
        ws.column_dimensions[get_column_letter(TREE_START + i)].width = 4.2

    # ── Row heights ───────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 20
    for d in range(max_d + 1):
        er = row_for_depth(d)
        for r in range(er, er + NODE_H):
            ws.row_dimensions[r].height = 20
        if d < max_d:
            conn_heights = [5, 7, 5, 5]  # cr_v, cr_h, cr_d, spacer
            for i, h in enumerate(conn_heights):
                ws.row_dimensions[er + NODE_H + i].height = h

    # ── Title row ─────────────────────────────────────────────────────────────
    total_cols = TREE_START + total_tree_cols + 2
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=total_cols)
    tc = ws.cell(1, 1)
    tc.value = 'แผนผังลำดับวงศ์ตระกูล'
    tc.font = Font(bold=True, size=18, color='FFFFFF', name='Sarabun')
    tc.fill = _fill(TITLE_FILL)
    tc.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=total_cols)
    sc = ws.cell(2, 1)
    sc.value = ('สีเข้ม = ต้นตระกูล  →  สีอ่อน = รุ่นหลัง   '
                '♥ = คู่สมรส   คลิกที่ช่องเพื่อดูรายละเอียด')
    sc.font = Font(italic=True, size=11, color='888888')
    sc.alignment = Alignment(horizontal='center', vertical='center')
    sc.fill = _fill('F8FAF9')

    # ── Sidebar labels ────────────────────────────────────────────────────────
    for d in range(max_d + 1):
        write_sidebar_label(ws, d, gen_years)

    # ── Draw nodes ────────────────────────────────────────────────────────────
    all_nodes: list = []
    for tree in trees:
        flatten_tree(tree, all_nodes)
    for node in all_nodes:
        if node['id'] in all_pos:
            write_node_card(ws, node, all_pos[node['id']])

    # ── Draw connectors ───────────────────────────────────────────────────────
    for tree in trees:
        draw_connectors(ws, tree, all_pos)

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = f'D3'


# ─── Sheet 2: Full member table ───────────────────────────────────────────────

def build_sheet_members(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet('ข้อมูลสมาชิกทั้งหมด')
    ws.sheet_view.showGridLines = False

    headers = [
        'ID', 'ชื่อ', 'นามสกุล', 'ชื่อเล่น', 'เพศ',
        'วันเกิด', 'วันเสียชีวิต', 'ภูมิลำเนา', 'อาชีพ',
        'บิดา-มารดา', 'คู่สมรส', 'บุตร-ธิดา',
        'บันทึกโดย', 'แก้ไขล่าสุดโดย', 'วันที่แก้ไข', 'หมายเหตุ',
    ]
    widths = [5, 14, 14, 10, 7, 18, 18, 16, 16, 26, 26, 30, 14, 14, 16, 20]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(1, 1)
    t.value = 'รายชื่อสมาชิกทั้งหมด'
    t.font = Font(bold=True, size=14, color='FFFFFF', name='Sarabun')
    t.fill = _fill(TITLE_FILL)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Header row
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(2, col, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=11, name='Sarabun')
        cell.fill = _fill(HEADER_FILL)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _border('555555')
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22

    members = list_members()
    alt = _fill('F4FBF6')

    for i, m in enumerate(members):
        row = i + 3
        fill = _fill('FFFFFF') if i % 2 == 0 else alt

        with get_conn() as conn:
            parents = conn.execute(
                "SELECT first_name||' '||last_name FROM members "
                "WHERE id IN (SELECT person1_id FROM relationships "
                "WHERE person2_id=? AND rel_type='parent-child')",
                (m['id'],)
            ).fetchall()
            spouses = conn.execute(
                "SELECT first_name||' '||last_name FROM members "
                "WHERE id IN ("
                "  SELECT CASE WHEN person1_id=? THEN person2_id ELSE person1_id END "
                "  FROM relationships WHERE rel_type='spouse' "
                "  AND (person1_id=? OR person2_id=?)"
                ")", (m['id'], m['id'], m['id'])
            ).fetchall()
            children = conn.execute(
                "SELECT first_name||' '||last_name FROM members "
                "WHERE id IN (SELECT person2_id FROM relationships "
                "WHERE person1_id=? AND rel_type='parent-child')",
                (m['id'],)
            ).fetchall()
            created_by_user = conn.execute(
                "SELECT full_name FROM users WHERE id=?", (m['created_by'],)
            ).fetchone() if m['created_by'] else None
            updated_by_user = conn.execute(
                "SELECT full_name FROM users WHERE id=?", (m['updated_by'],)
            ).fetchone() if m['updated_by'] else None

        row_vals = [
            m['id'],
            m['first_name'],
            m['last_name'],
            m['nickname'] or '',
            m['gender'] or '',
            display_date(m['birth_date'] or ''),
            display_date(m['death_date'] or ''),
            m['birth_place'] or '',
            m['occupation'] or '',
            ', '.join(r[0] for r in parents),
            ', '.join(r[0] for r in spouses),
            ', '.join(r[0] for r in children),
            created_by_user['full_name'] if created_by_user else '',
            updated_by_user['full_name'] if updated_by_user else '',
            (m['updated_at'] or '')[:16],
            m['notes'] or '',
        ]

        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row, col, value=val)
            cell.fill = fill
            cell.font = Font(size=10, name='Sarabun')
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = _border('DDDDDD')
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(headers))}{len(members)+2}'


# ─── Sheet 3: Legend / Instructions ──────────────────────────────────────────

def build_sheet_legend(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet('วิธีใช้งาน')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 65

    rows = [
        ('🌳  แผนผังลำดับวงศ์ตระกูล  —  คำแนะนำการใช้งาน', True,  TITLE_FILL, 'FFFFFF', 14, 32),
        ('',                                              False, 'FFFFFF',   '000000', 11, 8),
        ('📋  ไฟล์นี้มี 3 หน้า:',                         True,  'F0F7F4',   '1A472A', 12, 22),
        ('   • แผนผังตระกูล  —  แผนภูมิต้นไม้ top-down สีแยกตามรุ่น',
                                                          False, 'F0F7F4',   '333333', 11, 18),
        ('   • ข้อมูลสมาชิกทั้งหมด  —  ตารางข้อมูลครบถ้วน (มี filter)',
                                                          False, 'F0F7F4',   '333333', 11, 18),
        ('   • วิธีใช้งาน  —  หน้านี้',
                                                          False, 'F0F7F4',   '333333', 11, 18),
        ('',                                              False, 'FFFFFF',   '000000', 11, 8),
        ('🔑  วันที่ในระบบ:',                              True,  'FFF9EC',   '7D4E00', 12, 22),
        ('   • พ.ศ. XXXX  — ปีพุทธศักราช (เก็บเป็น BE:XXXX)',
                                                          False, 'FFF9EC',   '555555', 11, 18),
        ('   • ค.ศ. XXXX  — ปีคริสต์ศักราช (เก็บเป็น CE:XXXX)',
                                                          False, 'FFF9EC',   '555555', 11, 18),
        ('   • ไม่พบข้อมูล — ไม่ทราบวันที่ (เก็บเป็น UNKNOWN)',
                                                          False, 'FFF9EC',   '555555', 11, 18),
        ('',                                              False, 'FFFFFF',   '000000', 11, 8),
        ('🎨  สีตามรุ่น (Generation):',                   True,  'F0F7F4',   '1A472A', 12, 22),
    ]

    # Color legend rows
    for i, (bg, fg, _) in enumerate(GEN_PALETTE):
        rows.append((f'   รุ่นที่ {i+1}', True, bg, fg, 11, 20))

    rows += [
        ('',                                              False, 'FFFFFF',   '000000', 11, 8),
        ('🔧  การเพิ่ม/แก้ไขข้อมูล:',                    True,  'EBF5FB',   '154360', 12, 22),
        ('   เปิด Web UI: python web_ui.py  →  http://localhost:8888',
                                                          False, 'EBF5FB',   '1A5276', 11, 18),
        ('   หรือ CLI: python cli.py',
                                                          False, 'EBF5FB',   '1A5276', 11, 18),
        ('   หรือ export ใหม่: python export_excel.py',
                                                          False, 'EBF5FB',   '1A5276', 11, 18),
    ]

    for i, (text, bold, bg, fg, sz, ht) in enumerate(rows, 1):
        cell = ws.cell(i, 1, value=text)
        cell.font = Font(bold=bold, size=sz, color=fg, name='Sarabun')
        cell.fill = _fill(bg)
        cell.alignment = Alignment(vertical='center', indent=1)
        ws.row_dimensions[i].height = ht


# ─── Main export function ─────────────────────────────────────────────────────

def export_excel() -> Path:
    wb = openpyxl.Workbook()
    build_sheet_tree(wb)
    build_sheet_members(wb)
    build_sheet_legend(wb)
    wb.save(OUTPUT)
    builtins.print(f'[OK] Excel saved -> {OUTPUT}')
    return OUTPUT


if __name__ == '__main__':
    export_excel()
